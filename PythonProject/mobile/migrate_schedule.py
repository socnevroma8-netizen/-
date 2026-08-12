from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import delete

from mobile.database import Base, SessionLocal, engine
from mobile.db_models import Lesson


PROJECT_ROOT = Path(__file__).resolve().parent.parent

SCHEDULE_FILES = [
    PROJECT_ROOT / "mobile" / "data" / "schedule.xlsx",
    PROJECT_ROOT / "schedule.xlsx",
]


def normalize_text(value) -> str:
    return str(value or "").strip()


def find_schedule_file() -> Optional[Path]:
    for path in SCHEDULE_FILES:
        if path.exists():
            return path
    return None


def read_schedule_file(path: Path) -> tuple[list[dict], list[str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)

    sheet_name = (
        "По классам"
        if "По классам" in workbook.sheetnames
        else workbook.sheetnames[0]
    )
    worksheet = workbook[sheet_name]

    header_row = 3
    start_row = 5

    class_map = {}

    for column in range(4, worksheet.max_column + 1, 3):
        class_name = normalize_text(
            worksheet.cell(header_row, column).value
        )

        if class_name:
            class_map[class_name] = {
                "subject_col": column,
                "cabinet_col": column + 1,
                "teacher_col": column + 2,
            }

    rows = []
    current_day = ""

    for row_index in range(start_row, worksheet.max_row + 1):
        day_value = worksheet.cell(row_index, 1).value
        lesson_number = worksheet.cell(row_index, 2).value
        lesson_time = worksheet.cell(row_index, 3).value

        if day_value:
            current_day = normalize_text(day_value)

        if lesson_number in (None, ""):
            continue

        try:
            lesson_sort = int(
                float(str(lesson_number).replace(",", "."))
            )
        except (TypeError, ValueError):
            continue

        for class_name, columns in class_map.items():
            subject = normalize_text(
                worksheet.cell(
                    row_index,
                    columns["subject_col"],
                ).value
            )
            cabinet = normalize_text(
                worksheet.cell(
                    row_index,
                    columns["cabinet_col"],
                ).value
            )
            teacher = normalize_text(
                worksheet.cell(
                    row_index,
                    columns["teacher_col"],
                ).value
            )

            if not subject and not cabinet and not teacher:
                continue

            rows.append(
                {
                    "day": current_day,
                    "lesson_sort": lesson_sort,
                    "time": normalize_text(lesson_time),
                    "class_name": class_name,
                    "subject": subject,
                    "cabinet": cabinet,
                    "teacher": teacher,
                }
            )

    return rows, list(class_map)


def migrate_schedule(source_path: Optional[Path] = None) -> int:
    Base.metadata.create_all(bind=engine)

    path = source_path or find_schedule_file()

    if not path:
        raise FileNotFoundError(
            "Файл расписания не найден. "
            "Положи schedule.xlsx в mobile/data/"
        )

    rows, class_names = read_schedule_file(path)

    if not rows:
        raise ValueError(
            "В Excel-файле не найдено ни одной строки расписания."
        )

    lessons = []

    for row in rows:
        class_name = row["class_name"]
        day = row["day"]
        subject = row["subject"]

        if not class_name or not day or not subject:
            continue

        lessons.append(
            Lesson(
                grade=class_name,
                profile=None,
                day=day,
                lesson_number=row["lesson_sort"],
                lesson_time=row["time"] or None,
                subject=subject,
                teacher_name=row["teacher"] or None,
                cabinet=row["cabinet"] or None,
            )
        )

    if not lessons:
        raise ValueError(
            "После обработки не осталось валидных уроков."
        )

    with SessionLocal() as db:
        try:
            db.execute(delete(Lesson))
            db.add_all(lessons)
            db.commit()
        except Exception:
            db.rollback()
            raise

    print("Миграция расписания завершена")
    print(f"Источник: {path}")
    print(f"Строк в Excel: {len(rows)}")
    print(f"Уроков записано: {len(lessons)}")
    print(f"Классов найдено: {len(class_names)}")

    return len(lessons)


if __name__ == "__main__":
    migrate_schedule()