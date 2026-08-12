from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook

from kivy.lang import Builder
from kivy.metrics import dp
from kivy.uix.screenmanager import Screen

from kivymd.uix.button import MDFlatButton
from kivymd.uix.card import MDCard
from kivymd.uix.label import MDLabel

KV = """
<TransferScreen>:
    MDBoxLayout:
        orientation: "vertical"
        spacing: "12dp"
        padding: "12dp"

        MDBoxLayout:
            adaptive_height: True
            spacing: "8dp"

            MDRaisedButton:
                text: "Назад"
                size_hint_x: None
                width: "100dp"
                on_release: root.manager.current = "home"

            MDLabel:
                text: "Трансфер"
                halign: "center"
                bold: True
                adaptive_height: True

            Widget:

        MDLabel:
            id: subtitle_label
            text: ""
            halign: "center"
            theme_text_color: "Secondary"
            adaptive_height: True

        MDLabel:
            id: info_label
            text: ""
            halign: "center"
            theme_text_color: "Secondary"
            adaptive_height: True

        ScrollView:
            do_scroll_x: False
            size_hint_y: None
            height: "52dp"

            MDBoxLayout:
                id: filter_bar
                orientation: "horizontal"
                adaptive_width: True
                spacing: "8dp"
                padding: "4dp", 0

        ScrollView:
            do_scroll_x: False

            MDBoxLayout:
                id: content_box
                orientation: "vertical"
                size_hint_y: None
                height: self.minimum_height
                spacing: "10dp"
                padding: "4dp", "4dp", "4dp", "20dp"
"""

Builder.load_string(KV)


class TransferScreen(Screen):
    def __init__(self, **kwargs):
        self.selected_filter = "Все"
        self.filter_buttons = {}
        self.transfers_data = {}
        super().__init__(**kwargs)

    def on_kv_post(self, base_widget):
        self.transfers_data = self.load_transfers_from_excel()
        self.build_filters()

    def on_pre_enter(self, *args):
        self.ids.subtitle_label.text = self.build_subtitle()
        self.ids.info_label.text = self.build_info_text()
        self.highlight_selected_filter()
        self.render_transfers()

    def build_subtitle(self):
        today = datetime.now()
        weekdays = {
            0: "Понедельник",
            1: "Вторник",
            2: "Среда",
            3: "Четверг",
            4: "Пятница",
            5: "Суббота",
            6: "Воскресенье",
        }
        months = {
            1: "января",
            2: "февраля",
            3: "марта",
            4: "апреля",
            5: "мая",
            6: "июня",
            7: "июля",
            8: "августа",
            9: "сентября",
            10: "октября",
            11: "ноября",
            12: "декабря",
        }
        return f"{weekdays[today.weekday()]}, {today.day} {months[today.month]}"

    def build_info_text(self):
        now = datetime.now()
        weekday = now.weekday()
        evening_mode = "Вечер ПН-ЧТ" if weekday in (0, 1, 2, 3) else "Пятница"
        return f"Сегодня будний день — используется колонка «{evening_mode}». Обновлено: {now.strftime('%d.%m.%Y, %H:%M')}"

    def find_excel_path(self):
        base_dir = Path(__file__).resolve().parent
        candidates = [
            base_dir / "transfers.xlsx",
            base_dir.parent / "transfers.xlsx",
            base_dir.parent.parent / "transfers.xlsx",
            Path.cwd() / "transfers.xlsx",
        ]
        for path in candidates:
            if path.exists():
                return path
        return None

    def load_transfers_from_excel(self):
        path = self.find_excel_path()
        if not path:
            return {}

        wb = load_workbook(path, data_only=True)
        result = {}

        for ws in wb.worksheets:
            title = str(ws["B1"].value or ws["A1"].value or "").strip()
            m = re.search(r"ТРАНСФЕР\s*№\s*(\d+)", title, re.IGNORECASE)
            if not m:
                continue

            number = m.group(1)
            route_title = title.replace("\n", " ").strip()

            stops = []
            current_stop = None

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                values = [self.clean_cell(v) for v in row if self.clean_cell(v)]

                if not values:
                    continue

                time_candidate = values[0] if values else ""
                if self.is_time(time_candidate) and len(values) >= 2:
                    stop_name = values[1]
                    stops.append({
                        "name": stop_name,
                        "morning": time_candidate,
                        "evening": "",
                        "friday": "",
                    })
                    current_stop = stops[-1]
                    continue

                if current_stop:
                    line = " ".join(values)
                    if "вечер" in line.lower() or "пн-чт" in line.lower():
                        times = self.extract_times(line)
                        if times:
                            current_stop["evening"] = times[0]
                    elif "пят" in line.lower():
                        times = self.extract_times(line)
                        if times:
                            current_stop["friday"] = times[0]

            if stops:
                result[f"№{number}"] = [
                    {
                        "number": f"№{number}",
                        "title": route_title,
                        "route": route_title,
                        "stops": stops,
                    }
                ]

        if result:
            result = {"Все": [item for group in result.values() for item in group], **result}

        return result

    def clean_cell(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).replace("\xa0", " ").replace("\r", " ").strip()

    def is_time(self, text):
        text = str(text).strip()
        return bool(re.fullmatch(r"\d{1,2}[.:]\d{2}", text))

    def extract_times(self, text):
        return re.findall(r"\d{1,2}[.:]\d{2}", text)

    def build_filters(self):
        bar = self.ids.filter_bar
        bar.clear_widgets()
        self.filter_buttons = {}

        filters = list(self.transfers_data.keys()) if self.transfers_data else ["Все"]
        if "Все" not in filters:
            filters.insert(0, "Все")

        for name in filters:
            btn = MDFlatButton(
                text=name,
                on_release=lambda x, value=name: self.select_filter(value),
            )
            self.filter_buttons[name] = btn
            bar.add_widget(btn)

        self.highlight_selected_filter()

    def highlight_selected_filter(self):
        selected = getattr(self, "selected_filter", "Все")

        for name, btn in self.filter_buttons.items():
            if name == selected:
                btn.md_bg_color = (0.16, 0.47, 0.8, 1)
                btn.theme_text_color = "Custom"
                btn.text_color = (1, 1, 1, 1)
            else:
                btn.md_bg_color = (0, 0, 0, 0)
                btn.theme_text_color = "Primary"
                btn.text_color = (0, 0, 0, 1)

    def select_filter(self, value):
        self.selected_filter = value
        self.highlight_selected_filter()
        self.render_transfers()

    def render_transfers(self):
        box = self.ids.content_box
        box.clear_widgets()

        transfers = self.transfers_data.get(self.selected_filter, [])

        if not transfers:
            box.add_widget(
                MDLabel(
                    text="Нет данных по трансферам",
                    halign="center",
                    adaptive_height=True,
                )
            )
            return

        for transfer in transfers:
            box.add_widget(self.build_transfer_card(transfer))

    def build_transfer_card(self, transfer):
        card = MDCard(
            orientation="vertical",
            size_hint_y=None,
            padding=dp(14),
            spacing=dp(10),
            radius=[16, 16, 16, 16],
            elevation=2,
        )
        card.bind(minimum_height=card.setter("height"))

        count_stops = len(transfer["stops"])

        card.add_widget(
            MDLabel(
                text=transfer["title"],
                bold=True,
                adaptive_height=True,
            )
        )

        card.add_widget(
            MDLabel(
                text=f"Остановок: {count_stops}",
                theme_text_color="Secondary",
                adaptive_height=True,
            )
        )

        for i, stop in enumerate(transfer["stops"], start=1):
            stop_card = MDCard(
                orientation="vertical",
                size_hint_y=None,
                padding=dp(10),
                spacing=dp(4),
                radius=[12, 12, 12, 12],
                elevation=0,
                md_bg_color=(0.95, 0.95, 0.95, 1),
            )
            stop_card.bind(minimum_height=stop_card.setter("height"))

            stop_card.add_widget(
                MDLabel(
                    text=f"{i}. {stop['name']}",
                    bold=True,
                    adaptive_height=True,
                )
            )
            stop_card.add_widget(MDLabel(text=f"Утро: {stop.get('morning', '')}", adaptive_height=True))
            if stop.get("evening"):
                stop_card.add_widget(MDLabel(text=f"Вечер: {stop['evening']}", adaptive_height=True))
            if stop.get("friday"):
                stop_card.add_widget(MDLabel(text=f"Пятница: {stop['friday']}", adaptive_height=True))

            card.add_widget(stop_card)

        return card