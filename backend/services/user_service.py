# backend/services/user_service.py
from pathlib import Path
from typing import Dict, Optional

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BASE_DIR.parent
DATA_DIR = PROJECT_ROOT / "mobile" / "data"
USERS_XLSX = DATA_DIR / "users.xlsx"


class UserService:
    def __init__(self) -> None:
        self.users: Dict[str, dict] = {}
        self._load_users()

    def _load_users(self) -> None:
        self.users = {}

        if not USERS_XLSX.exists():
            print(f"[user_service] users.xlsx not found: {USERS_XLSX}")
            return

        wb = openpyxl.load_workbook(USERS_XLSX, data_only=True)

        # --- Ученики ---
        if "Ученики" in wb.sheetnames:
            ws = wb["Ученики"]
            headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[1]
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                login = str(data.get("Логин") or "").strip()
                password = str(data.get("Пароль") or "").strip()

                if not login or not password:
                    continue

                self.users[login] = {
                    "id": len(self.users) + 1,
                    "fio": str(data.get("ФИО") or "").strip(),
                    "role": "student",
                    "klass": str(data.get("Класс") or "").strip(),
                    "profile": str(data.get("Профиль") or "").strip(),
                    "password": password,
                    "is_admin": str(data.get("Админ") or "").strip().lower() == "админ",
                    "teacher_subject": None,
                    "teacher_cabinet": None,
                    "teacher_classroom_teacher": None,
                }

        # --- Педагоги ---
        if "Педагоги" in wb.sheetnames:
            ws = wb["Педагоги"]
            headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[1]
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                login = str(data.get("Логин") or "").strip()
                password = str(data.get("Пароль") or "").strip()

                if not login or not password:
                    continue

                self.users[login] = {
                    "id": len(self.users) + 1,
                    "fio": str(data.get("ФИО") or "").strip(),
                    "role": "teacher",
                    "klass": None,
                    "profile": None,
                    "password": password,
                    "is_admin": str(data.get("Админ") or "").strip().lower() == "админ",
                    "teacher_subject": str(data.get("Предмет") or "").strip(),
                    "teacher_cabinet": str(data.get("Кабинет") or "").strip(),
                    "teacher_classroom_teacher": str(data.get("Классное руководство") or "").strip(),
                }

        print("[user_service] loaded users:", list(self.users.keys()))

    def reload(self) -> None:
        self._load_users()
    def login(self, login: str, password: str) -> Optional[dict]:
        user = self.users.get(login)
        if not user:
            return None

        if user["password"] != password:
            return None

        # Возвращаем полную структуру, UI потом нормализует под себя
        return {
            "id": user["id"],
            "fio": user["fio"],
            "role": user["role"],
            "is_admin": user["is_admin"],
            "klass": user["klass"],
            "profile": user["profile"],
            "teacher_subject": user["teacher_subject"],
            "teacher_cabinet": user["teacher_cabinet"],
            "teacher_classroom_teacher": user["teacher_classroom_teacher"],
        }

    def create_user(
        self,
        fio: str,
        role: str,         # "student" или "teacher"
        login: str,
        password: str,
        klass: str = "",
        profile: str = "",
        subject: str = "",
        cabinet: str = "",
        classroom: str = "",
        admin: bool = False,
    ) -> dict:
        """
        Добавляет ученика или педагога в users.xlsx и обновляет self.users.
        Возвращает такую же структуру, как login().
        """

        if login in self.users:
            raise ValueError(f"Пользователь с логином '{login}' уже существует")

        if not USERS_XLSX.exists():
            raise ValueError(f"Файл users.xlsx не найден: {USERS_XLSX}")

        wb = openpyxl.load_workbook(USERS_XLSX, data_only=True)  # [web:652][web:677]

        if role == "student":
            sheet_name = "Ученики"
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Лист '{sheet_name}' не найден в users.xlsx")

            ws = wb[sheet_name]
            headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[1]
            ]

            row_values = []
            for h in headers:
                if h == "ФИО":
                    row_values.append(fio)
                elif h == "Класс":
                    row_values.append(klass)
                elif h == "Профиль":
                    row_values.append(profile)
                elif h == "Логин":
                    row_values.append(login)
                elif h == "Пароль":
                    row_values.append(password)
                elif h == "Админ":
                    row_values.append("админ" if admin else "")
                else:
                    row_values.append(None)

            ws.append(row_values)  # [web:671][web:678]

        elif role == "teacher":
            sheet_name = "Педагоги"
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Лист '{sheet_name}' не найден в users.xlsx")

            ws = wb[sheet_name]
            headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[1]
            ]

            row_values = []
            for h in headers:
                if h == "ФИО":
                    row_values.append(fio)
                elif h == "Предмет":
                    row_values.append(subject)
                elif h == "Кабинет":
                    row_values.append(cabinet)
                elif h == "Классное руководство":
                    row_values.append(classroom)
                elif h == "Логин":
                    row_values.append(login)
                elif h == "Пароль":
                    row_values.append(password)
                elif h == "Админ":
                    row_values.append("админ" if admin else "")
                else:
                    row_values.append(None)

            ws.append(row_values)  # [web:671][web:678]

        else:
            raise ValueError("role должен быть 'student' или 'teacher'")

        wb.save(USERS_XLSX)  # [web:666][web:677]

        # обновляем кэш
        self._load_users()

        created = self.users.get(login)
        if not created:
            raise ValueError("Не удалось перечитать созданного пользователя")

        return {
            "id": created["id"],
            "fio": created["fio"],
            "role": created["role"],
            "is_admin": created["is_admin"],
            "klass": created["klass"],
            "profile": created["profile"],
            "teacher_subject": created["teacher_subject"],
            "teacher_cabinet": created["teacher_cabinet"],
            "teacher_classroom_teacher": created["teacher_classroom_teacher"],
        }