from datetime import datetime
from pathlib import Path
import re

from fastapi import APIRouter, Request, Form, Depends
from fastapi.responses import RedirectResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import text
from openpyxl import load_workbook

from .models.db import SessionLocal

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
router = APIRouter(prefix="/web", tags=["web"])
DAY_ORDER = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def current_user(request: Request, db):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    row = db.execute(text("SELECT id, full_name, username, role, is_admin, grade, profile FROM users WHERE id = :id"), {"id": user_id}).mappings().first()
    return dict(row) if row else None


def class_key_for(user):
    return f'{user["grade"]} / {user["profile"]}' if user.get("profile") else user["grade"]


def redirect_if_anonymous(request: Request, db):
    user = current_user(request, db)
    return (None, RedirectResponse(url="/web/login", status_code=302)) if not user else (user, None)


def lessons_for_class(db, class_key):
    return db.execute(text("SELECT day, lesson_number, lesson_time, subject, teacher_name, cabinet FROM lessons WHERE grade = :grade ORDER BY lesson_number"), {"grade": class_key}).mappings().all()


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").replace("\r", " ").strip()


def extract_times(value):
    return re.findall(r"\d{1,2}[.:]\d{2}", value)


def find_transfers_file():
    candidates = [BASE_DIR / "mobile" / "data" / "transfers.xlsx", BASE_DIR / "transfers.xlsx", BASE_DIR.parent / "transfers.xlsx"]
    return next((path for path in candidates if path.exists()), None)


def load_transfers():
    path = find_transfers_file()
    if not path:
        return []
    workbook = load_workbook(path, data_only=True, read_only=True)
    result = []
    for worksheet in workbook.worksheets:
        title = clean_cell(worksheet["B1"].value or worksheet["A1"].value)
        if not re.search(r"ТРАНСФЕР\s*№\s*\d+", title, re.IGNORECASE):
            continue
        stops = []
        current_stop = None
        for row in worksheet.iter_rows(values_only=True):
            values = [clean_cell(value) for value in row if clean_cell(value)]
            if not values:
                continue
            if re.fullmatch(r"\d{1,2}[.:]\d{2}", values[0]) and len(values) >= 2:
                current_stop = {"name": values[1], "morning": values[0], "evening": "", "friday": ""}
                stops.append(current_stop)
                continue
            if current_stop:
                line = " ".join(values)
                times = extract_times(line)
                if times and ("вечер" in line.lower() or "пн-чт" in line.lower()):
                    current_stop["evening"] = times[0]
                elif times and "пят" in line.lower():
                    current_stop["friday"] = times[0]
        if stops:
            result.append({"number": title, "title": title, "stops": stops})
    return result


@router.get("/", response_class=HTMLResponse)
def web_root(request: Request, db=Depends(get_db)):
    user = current_user(request, db)
    return RedirectResponse(url="/web/home" if user else "/web/login", status_code=302)


@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request=request, name="web_login.html", context={"error": None})


@router.post("/login", response_class=HTMLResponse)
def login_submit(request: Request, username: str = Form(...), password: str = Form(...), db=Depends(get_db)):
    row = db.execute(text("SELECT id, full_name, username, password, role, is_admin, grade, profile FROM users WHERE username = :username"), {"username": username}).mappings().first()
    if not row or row["password"] != password:
        return templates.TemplateResponse(request=request, name="web_login.html", context={"error": "Неверный логин или пароль"}, status_code=401)
    request.session["user_id"] = row["id"]
    request.session["username"] = row["username"]
    request.session["full_name"] = row["full_name"]
    request.session["is_admin"] = row["is_admin"]
    return RedirectResponse(url="/web/home", status_code=302)


@router.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/web/login", status_code=302)


@router.get("/home", response_class=HTMLResponse)
def home_page(request: Request, db=Depends(get_db)):
    user, redirect = redirect_if_anonymous(request, db)
    if redirect:
        return redirect
    today = datetime.now()
    day_name = DAY_ORDER[today.weekday()]
    lessons = [dict(row) for row in lessons_for_class(db, class_key_for(user)) if row["day"] == day_name]
    return templates.TemplateResponse(request=request, name="web_home.html", context={"user": user, "lessons": lessons, "today_label": today.strftime("%d.%m.%Y")})


@router.get("/schedule", response_class=HTMLResponse)
def schedule_page(request: Request, db=Depends(get_db)):
    user, redirect = redirect_if_anonymous(request, db)
    if redirect:
        return redirect
    rows = lessons_for_class(db, class_key_for(user))
    days = {}
    for row in rows:
        days.setdefault(row["day"], []).append(dict(row))
    ordered_days = [(day, days[day]) for day in DAY_ORDER if day in days]
    return templates.TemplateResponse(request=request, name="web_schedule.html", context={"user": user, "days": ordered_days, "updated_at": datetime.now().strftime("%d.%m.%Y %H:%M")})


@router.get("/transfers", response_class=HTMLResponse)
def transfers_page(request: Request, db=Depends(get_db)):
    user, redirect = redirect_if_anonymous(request, db)
    if redirect:
        return redirect
    return templates.TemplateResponse(request=request, name="web_section.html", context={"user": user, "title": "Трансферы", "message": "", "transfers": load_transfers()})


@router.get("/{section}", response_class=HTMLResponse)
def section_page(section: str, request: Request, db=Depends(get_db)):
    user, redirect = redirect_if_anonymous(request, db)
    if redirect:
        return redirect
    sections = {"users": ("База пользователей", "Раздел пользователей будет доступен администраторам."), "reminders": ("Напоминания", "Здесь будут отображаться напоминания.")}
    if section not in sections:
        return HTMLResponse("Раздел не найден", status_code=404)
    title, message = sections[section]
    if section == "users" and not user.get("is_admin"):
        return HTMLResponse("Доступ разрешён только администраторам", status_code=403)
    return templates.TemplateResponse(request=request, name="web_section.html", context={"user": user, "title": title, "message": message})
