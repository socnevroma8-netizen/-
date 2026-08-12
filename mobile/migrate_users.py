from pathlib import Path

import openpyxl
from sqlalchemy import select

from mobile.database import Base, SessionLocal, engine
from mobile.db_models import Teacher, User


PROJECT_ROOT = Path(__file__).resolve().parent.parent
USERS_XLSX = PROJECT_ROOT / "mobile" / "data" / "users.xlsx"


def value(data: dict, column: str) -> str:
    return str(data.get(column) or "").strip()


def is_admin(data: dict) -> bool:
    return value(data, "Админ").lower() == "админ"


def read_sheet_rows(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        print(f"[skip] В users.xlsx нет листа «{sheet_name}»")
        return []

    worksheet = workbook[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]

    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def user_exists(db, username: str) -> bool:
    return db.scalar(select(User.id).where(User.username == username)) is not None


def teacher_exists(db, full_name: str) -> bool:
    return db.scalar(select(Teacher.id).where(Teacher.full_name == full_name)) is not None


def migrate_students(db, rows: list[dict]) -> tuple[int, int]:
    added = 0
    skipped = 0

    for row in rows:
        username = value(row, "Логин")
        password = value(row, "Пароль")
        full_name = value(row, "ФИО")

        if not username or not password or not full_name:
            print(f"[skip student] Неполная строка: логин={username!r}, ФИО={full_name!r}")
            skipped += 1
            continue

        if user_exists(db, username):
            skipped += 1
            continue

        db.add(
            User(
                full_name=full_name,
                username=username,
                password=password,
                role="student",
                is_admin=is_admin(row),
                grade=value(row, "Класс") or None,
                profile=value(row, "Профиль") or None,
            )
        )
        added += 1

    return added, skipped


def migrate_teachers(db, rows: list[dict]) -> tuple[int, int, int]:
    users_added = 0
    teachers_added = 0
    skipped = 0

    for row in rows:
        username = value(row, "Логин")
        password = value(row, "Пароль")
        full_name = value(row, "ФИО")

        if not username or not password or not full_name:
            print(f"[skip teacher] Неполная строка: логин={username!r}, ФИО={full_name!r}")
            skipped += 1
            continue

        if not user_exists(db, username):
            db.add(
                User(
                    full_name=full_name,
                    username=username,
                    password=password,
                    role="teacher",
                    is_admin=is_admin(row),
                    grade=None,
                    profile=None,
                )
            )
            users_added += 1
        else:
            skipped += 1

        if not teacher_exists(db, full_name):
            db.add(
                Teacher(
                    full_name=full_name,
                    subject=value(row, "Предмет") or None,
                    cabinet=value(row, "Кабинет") or None,
                    homeroom=value(row, "Классное руководство") or None,
                )
            )
            teachers_added += 1

    return users_added, teachers_added, skipped


def main():
    if not USERS_XLSX.exists():
        raise FileNotFoundError(f"Не найден users.xlsx: {USERS_XLSX}")

    Base.metadata.create_all(bind=engine)
    workbook = openpyxl.load_workbook(USERS_XLSX, data_only=True)

    student_rows = read_sheet_rows(workbook, "Ученики")
    teacher_rows = read_sheet_rows(workbook, "Педагоги")

    with SessionLocal() as db:
        try:
            students_added, students_skipped = migrate_students(db, student_rows)
            teacher_users_added, teachers_added, teachers_skipped = migrate_teachers(db, teacher_rows)
            db.commit()
        except Exception:
            db.rollback()
            raise

    print("\nМиграция завершена")
    print(f"Ученики: добавлено {students_added}, пропущено {students_skipped}")
    print(f"Пользователи-педагоги: добавлено {teacher_users_added}")
    print(f"Педагоги: добавлено {teachers_added}, пропущено пользователей {teachers_skipped}")


if __name__ == "__main__":
    main()